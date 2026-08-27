
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
import zipfile
import io
import re
import pandas as pd
import base64
from datetime import datetime

st.set_page_config(
    page_title="Pushpak Kumar | GSTR-2B Consolidator",
    page_icon="icon128.png",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ---------------- PUSHPAK KUMAR BRANDING ----------------
_logo = Path(__file__).parent / "icon128.png"
_logo_b64 = base64.b64encode(_logo.read_bytes()).decode("utf-8") if _logo.exists() else ""

st.markdown("""
<style>
.stApp { background:#f5f7fb; }
.block-container { max-width:1180px; padding-top:1.2rem; padding-bottom:3rem; }
#MainMenu, footer { visibility:hidden; }

.pk-brand {
    background:#071a33; border-radius:18px; padding:16px 22px;
    display:flex; align-items:center; justify-content:space-between;
    margin-bottom:22px; box-shadow:0 8px 28px rgba(7,26,51,.12);
}
.pk-left { display:flex; align-items:center; gap:14px; }
.pk-logo { width:58px; height:58px; object-fit:contain; background:#000;
           border-radius:10px; padding:3px; }
.pk-name { color:#fff; font-size:20px; font-weight:750; }
.pk-sub { color:#b9c8dc; font-size:12px; margin-top:3px; }
.pk-site { color:#dce8f7; font-size:13px; font-weight:650; }

.pk-hero {
    background:#fff; border:1px solid #e5eaf2; border-radius:20px;
    padding:30px 34px; margin-bottom:18px;
    box-shadow:0 5px 24px rgba(15,34,58,.05);
}
.pk-eyebrow { color:#2463a6; font-size:11px; font-weight:800;
              letter-spacing:1.5px; text-transform:uppercase; }
.pk-title { color:#071a33; font-size:34px; font-weight:800;
            line-height:1.15; margin-top:7px; }
.pk-desc { color:#64748b; font-size:14px; margin-top:8px; }

.pk-features { display:grid; grid-template-columns:repeat(4,1fr);
               gap:12px; margin:15px 0 22px; }
.pk-card { background:#fff; border:1px solid #e5eaf2; border-radius:14px;
           padding:15px; min-height:85px; }
.pk-card b { color:#142a45; font-size:13px; }
.pk-card span { display:block; color:#718096; font-size:11px;
                line-height:1.45; margin-top:5px; }

[data-testid="stFileUploader"] {
    background:#fff; border:2px dashed #aebed2;
    border-radius:18px; padding:10px;
}
.stButton > button, .stDownloadButton > button {
    border-radius:10px; font-weight:700; min-height:46px;
}
.pk-section { color:#071a33; font-size:18px; font-weight:750;
              margin:23px 0 9px; }
.pk-footer { margin-top:32px; padding-top:16px; border-top:1px solid #dde4ee;
             color:#77869a; font-size:11px; text-align:center; }

@media(max-width:850px) {
 .pk-features { grid-template-columns:repeat(2,1fr); }
 .pk-site { display:none; }
 .pk-title { font-size:28px; }
}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="pk-brand">
  <div class="pk-left">
    <img class="pk-logo" src="data:image/png;base64,{_logo_b64}">
    <div>
      <div class="pk-name">Pushpak Kumar</div>
      <div class="pk-sub">GST Automation Toolkit</div>
    </div>
  </div>
  <div class="pk-site">pushpakkumar.com</div>
</div>
""", unsafe_allow_html=True)
# --------------------------------------------------------

# ================================================================
# FIXED GSTR-2B OUTPUT STRUCTURE
# ================================================================
SUMMARY_SHEETS = [
    "ITC Available",
    "ITC not available",
    "ITC Reversal",
    "ITC Rejected",
]

TRANSACTION_MAP = {
    "B2B": ["B2B", "B2B (2)"],
    "B2BA": ["B2BA"],
    "B2B-CDNR": ["B2B-CDNR"],
    "B2B-CDNRA": ["B2B-CDNRA"],
    "ECO": ["ECO"],
    "ECOA": ["ECOA"],
    "ISD": ["ISD"],
    "ISDA": ["ISDA"],
    "IMPG": ["IMPG"],
    "IMPGA": ["IMPGA"],
    "IMPGSEZ": ["IMPGSEZ"],
    "IMPGSEZA": ["IMPGSEZA"],
    "B2B Reversal": ["B2B (ITC Reversal)", "B2B Reversal"],
    "B2BA Reversal": ["B2BA (ITC Reversal)", "B2BA Reversal"],
    "B2B DNR": ["B2B-DNR", "B2B DNR"],
    "B2B DNRA": ["B2B-DNRA", "B2B DNRA"],
}
TRANSACTION_SHEETS = list(TRANSACTION_MAP.keys())

SUMMARY_NUMERIC_HEADERS = [
    "integratedtax", "centraltax", "stateuttax", "cess",
    "integratedtax", "centraltax", "stateut tax"
]

def norm(v):
    return re.sub(r"[^a-z0-9]+", "", str(v or "").strip().lower())

def clean(v):
    return "" if v is None else re.sub(r"\s+", " ", str(v).strip())

def month_from_filename(filename):
    s = Path(filename).stem

    # Portal naming: 012026_... = Jan-26
    m = re.search(r"(?<!\d)(0[1-9]|1[0-2])(\d{4})(?!\d)", s)
    if m:
        return pd.Timestamp(int(m.group(2)), int(m.group(1)), 1).strftime("%b-%y")

    m = re.search(r"(?<!\d)(0[1-9]|1[0-2])[-_](20\d{2})(?!\d)", s)
    if m:
        return pd.Timestamp(int(m.group(2)), int(m.group(1)), 1).strftime("%b-%y")

    m = re.search(r"(?<!\d)(20\d{2})[-_](0[1-9]|1[0-2])(?!\d)", s)
    if m:
        return pd.Timestamp(int(m.group(1)), int(m.group(2)), 1).strftime("%b-%y")

    return "Unknown"

def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def numeric_columns(ws, header_row):
    """
    Detect numeric tax/amount columns from the portal header.
    We aggregate ONLY numeric columns. Text/layout is never added.
    """
    cols = []
    for c in range(1, ws.max_column + 1):
        pieces = []
        for r in range(max(1, header_row-1), header_row+1):
            pieces.append(norm(ws.cell(r,c).value))
        h = "".join(pieces)
        if any(x in h for x in ["integratedtax", "centraltax", "stateuttax", "cess", "taxablevalue", "invoicevalue", "notevalue"]):
            cols.append(c)
    return cols

def find_summary_header(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = {norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)}
        if "sno" in vals and "heading" in vals:
            return r
    return 6

def summary_key(ws, row):
    """
    Stable row identity for summary aggregation.
    Prefer S.no + Heading + GSTR-3B table.
    """
    vals = [clean(ws.cell(row,c).value) for c in range(1, min(4, ws.max_column)+1)]
    return tuple(vals)

def copy_sheet_top(src, dst, max_row):
    # Copy widths
    for c in range(1, src.max_column+1):
        letter = get_column_letter(c)
        dst.column_dimensions[get_column_letter(c+1)].width = src.column_dimensions[letter].width

    # Copy rows/cells
    for r in range(1, max_row+1):
        dst.row_dimensions[r].height = src.row_dimensions[r].height
        for c in range(1, src.max_column+1):
            s = src.cell(r,c)
            d = dst.cell(r,c+1)
            d.value = s.value
            if s.has_style:
                d._style = copy(s._style)
            d.number_format = s.number_format
            d.font = copy(s.font)
            d.fill = copy(s.fill)
            d.border = copy(s.border)
            d.alignment = copy(s.alignment)
            d.protection = copy(s.protection)

    # Shift merged ranges one column
    for merged in src.merged_cells.ranges:
        if merged.max_row <= max_row:
            dst.merge_cells(
                start_row=merged.min_row, start_column=merged.min_col+1,
                end_row=merged.max_row, end_column=merged.max_col+1
            )

def add_month_column(ws, header_start, header_end):
    if header_end > header_start:
        ws.merge_cells(
            start_row=header_start, start_column=1,
            end_row=header_end, end_column=1
        )
    cell = ws.cell(header_start, 1, "GSTR-2B Month")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 16

def transaction_header_end(ws):
    """
    Find the actual last header row from the portal.
    GSTR-2B detail sheets have the tax column labels in the second header row.
    """
    for r in range(1, min(ws.max_row, 15)+1):
        vals = [norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        if any("integratedtax" in v for v in vals):
            return r
    return 6

def repeated_transaction_header(vals):
    s = " ".join(norm(v) for v in vals if v is not None)
    hits = 0
    for k in [
        "invoicenumber", "invoicetype", "invoicedate",
        "notenumber", "notetype", "notedate",
        "gstinofsupplier", "tradename",
        "integratedtax", "centraltax", "stateuttax",
        "billofentrynumber", "portcode"
    ]:
        if k in s:
            hits += 1
    return hits >= 3

def transaction_rows(ws, header_end):
    result = []
    for r in range(header_end+1, ws.max_row+1):
        vals = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]

        if not any(v is not None and str(v).strip() for v in vals):
            continue

        if repeated_transaction_header(vals):
            continue

        result.append(vals)
    return result

def actual_matches(wb, target):
    aliases = {norm(x) for x in TRANSACTION_MAP[target]}
    return [s for s in wb.sheetnames if norm(s) in aliases]

# ================================================================
# TRANSACTION SHEETS
# ================================================================
def combine_transaction_sheet(files, target, out_wb):
    template_wb = None
    template_name = None

    for filename, raw in files:
        wb = load_workbook(io.BytesIO(raw), data_only=False, read_only=False)
        matches = actual_matches(wb, target)
        if matches:
            template_wb = wb
            template_name = matches[0]
            break
        wb.close()

    if template_wb is None:
        ws = out_wb.create_sheet(target)
        ws["A1"] = "GSTR-2B Month"
        return 0, "No source sheet"

    template_ws = template_wb[template_name]
    header_end = transaction_header_end(template_ws)
    header_start = max(1, header_end-1)

    ws = out_wb.create_sheet(target)

    # KEEP THE ORIGINAL PORTAL FORMAT/HEADERS.
    copy_sheet_top(template_ws, ws, header_end)
    add_month_column(ws, header_start, header_end)

    next_row = header_end + 1
    total_rows = 0

    for filename, raw in files:
        month = month_from_filename(filename)

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        matches = actual_matches(wb, target)

        for actual in matches:
            src = wb[actual]
            src_header_end = transaction_header_end(src)

            for vals in transaction_rows(src, src_header_end):
                # Source columns are copied exactly under the permanent template.
                width = template_ws.max_column
                vals = list(vals[:width])
                if len(vals) < width:
                    vals += [None] * (width-len(vals))

                ws.cell(next_row, 1, month)
                ws.cell(next_row, 1).number_format = "@"

                for c, value in enumerate(vals, start=2):
                    ws.cell(next_row, c, value)

                next_row += 1
                total_rows += 1

        wb.close()

    template_wb.close()

    ws.freeze_panes = f"A{header_end+1}"
    if ws.max_row >= header_end+1:
        ws.auto_filter.ref = f"A{header_end}:{get_column_letter(ws.max_column)}{ws.max_row}"

    return total_rows, "OK"

# ================================================================
# SUMMARY SHEETS — TOTALS, NOT MONTH-WISE
# ================================================================
def find_summary_header(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = {norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)}
        if "sno" in vals and "heading" in vals:
            return r
    return 6

def summary_amount_columns(ws, header_row):
    """
    GSTR-2B summary amount columns are Integrated Tax, Central Tax,
    State/UT Tax and Cess. Only these numeric figure columns are added.
    """
    cols = []
    for c in range(1, ws.max_column+1):
        h1 = norm(ws.cell(header_row, c).value)
        h2 = norm(ws.cell(header_row-1, c).value) if header_row > 1 else ""
        h = h1 + h2
        if any(x in h for x in [
            "integratedtax", "centraltax", "stateuttax", "cess"
        ]):
            cols.append(c)
    return cols

def summary_row_key(ws, row, header_row):
    """
    Match a summary row by its visible portal labels.
    This is safer than relying only on row number if a portal export
    changes slightly between periods.
    """
    vals = []
    for c in range(1, min(ws.max_column, 3)+1):
        vals.append(clean(ws.cell(row,c).value))

    # Include the first non-empty text cells beyond the first three if needed.
    # This helps distinguish detail rows with blank S.no.
    if all(v == "" for v in vals):
        for c in range(1, min(ws.max_column, 8)+1):
            v = clean(ws.cell(row,c).value)
            if v:
                vals.append(v)
                if len(vals) >= 4:
                    break

    return tuple(vals)

def combine_summary_sheet(files, sheet_name, out_wb):
    """
    IMPORTANT:
    Summary sheets are NOT month-wise.

    One month's original GSTR-2B summary layout is retained exactly.
    The Integrated Tax / Central Tax / State-UT Tax / Cess figures from
    every uploaded month are added into the SAME corresponding cells.

    Example:
        Jan cell = 100
        Feb cell = 200
        Mar cell = 300
        Final cell = 600

    No GSTR-2B Month column is added to the summary sheets.
    """
    template_wb = None
    template_ws = None

    for filename, raw in files:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        if sheet_name in wb.sheetnames:
            template_wb = wb
            template_ws = wb[sheet_name]
            break
        wb.close()

    if template_ws is None:
        ws = out_wb.create_sheet(sheet_name)
        ws["A1"] = sheet_name
        return 0, "No source sheet"

    hrow = find_summary_header(template_ws)

    # Copy the ORIGINAL summary page without inserting any extra column.
    ws = out_wb.create_sheet(sheet_name)

    for r in range(1, template_ws.max_row+1):
        ws.row_dimensions[r].height = template_ws.row_dimensions[r].height
        for c in range(1, template_ws.max_column+1):
            s = template_ws.cell(r,c)
            d = ws.cell(r,c)
            d.value = s.value
            if s.has_style:
                d._style = copy(s._style)
            d.number_format = s.number_format
            d.font = copy(s.font)
            d.fill = copy(s.fill)
            d.border = copy(s.border)
            d.alignment = copy(s.alignment)
            d.protection = copy(s.protection)

    for c in range(1, template_ws.max_column+1):
        ws.column_dimensions[get_column_letter(c)].width = template_ws.column_dimensions[get_column_letter(c)].width

    for merged in template_ws.merged_cells.ranges:
        ws.merge_cells(
            start_row=merged.min_row, start_column=merged.min_col,
            end_row=merged.max_row, end_column=merged.max_col
        )

    amount_cols = summary_amount_columns(template_ws, hrow)

    # Template row keys.
    template_rows = {}
    for r in range(hrow+1, template_ws.max_row+1):
        key = summary_row_key(template_ws, r, hrow)
        if key and any(key):
            template_rows[key] = r

    numeric_cells_added = 0
    source_months = []

    for filename, raw in files:
        month = month_from_filename(filename)
        source_months.append(month)

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            continue

        src = wb[sheet_name]
        src_hrow = find_summary_header(src)
        src_amount_cols = summary_amount_columns(src, src_hrow)

        # Header-text -> source column mapping.
        src_col_map = {}
        for c in src_amount_cols:
            h = norm(src.cell(src_hrow,c).value)
            if h:
                src_col_map[h] = c

        # Source row mapping.
        source_rows = {}
        for r in range(src_hrow+1, src.max_row+1):
            key = summary_row_key(src, r, src_hrow)
            if key and any(key) and key not in source_rows:
                source_rows[key] = r

        for key, out_r in template_rows.items():
            src_r = source_rows.get(key)
            if src_r is None:
                continue

            for out_c in amount_cols:
                header = norm(template_ws.cell(hrow,out_c).value)
                src_c = src_col_map.get(header)

                if src_c is None:
                    continue

                value = src.cell(src_r, src_c).value

                if is_number(value):
                    current = ws.cell(out_r,out_c).value
                    if not is_number(current):
                        current = 0
                    ws.cell(out_r,out_c).value = current + value
                    numeric_cells_added += 1

        wb.close()

    template_wb.close()

    # Keep the original portal header/filter exactly.
    ws.freeze_panes = template_ws.freeze_panes
    return numeric_cells_added, "TOTALLED"

# ================================================================
# READ ME / VALIDATION
# ================================================================
def create_readme(files, summary_counts, transaction_counts, log, wb):
    """Create a professional branded Excel landing page."""
    ws = wb.create_sheet("Read me")

    ws.sheet_view.showGridLines = False

    # Column layout
    widths = {"A":4, "B":20, "C":20, "D":20, "E":20, "F":20, "G":20, "H":20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Brand colours
    navy = "071A33"
    blue = "1F4E78"
    soft = "EAF0F7"
    grey = "617086"
    green = "2F6B4F"

    # Logo
    logo_path = Path(__file__).parent / "icon128.png"
    if logo_path.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(logo_path))
            img.width = 68
            img.height = 68
            ws.add_image(img, "B2")
        except Exception:
            pass

    # Brand bar
    for row in range(2, 4):
        for col in range(2, 9):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=navy)

    ws.merge_cells("C2:F2")
    ws["C2"] = "PUSHPAK KUMAR"
    ws["C2"].font = Font(size=19, bold=True, color="FFFFFF")
    ws["C2"].fill = PatternFill("solid", fgColor=navy)
    ws["C2"].alignment = Alignment(vertical="center")

    ws.merge_cells("C3:F3")
    ws["C3"] = "GST AUTOMATION TOOLKIT"
    ws["C3"].font = Font(size=10, bold=True, color="DCE8F7")
    ws["C3"].fill = PatternFill("solid", fgColor=navy)

    ws.merge_cells("G2:H3")
    ws["G2"] = "pushpakkumar.com"
    ws["G2"].font = Font(size=11, bold=True, color="FFFFFF", underline="single")
    ws["G2"].fill = PatternFill("solid", fgColor=navy)
    ws["G2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G2"].hyperlink = "https://pushpakkumar.com"

    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 25

    # Main title
    ws.merge_cells("B5:H5")
    ws["B5"] = "GSTR-2B CONSOLIDATED WORKBOOK"
    ws["B5"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["B5"].fill = PatternFill("solid", fgColor=blue)
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 30

    periods = ", ".join(month_from_filename(n) for n,_ in files)
    ws.merge_cells("B6:H6")
    ws["B6"] = f"GST Automation Toolkit  •  Periods: {periods}"
    ws["B6"].font = Font(size=10, color=grey)
    ws["B6"].alignment = Alignment(horizontal="center")

    # KPI cards
    cards = [
        ("B8:C8", "B9:C10", "SOURCE FILES", len(files)),
        ("D8:E8", "D9:E10", "TRANSACTION ROWS", sum(transaction_counts.values())),
        ("F8:G8", "F9:G10", "PERIODS", len(set(month_from_filename(n) for n,_ in files))),
    ]
    for head_rng, value_rng, label, value in cards:
        ws.merge_cells(head_rng)
        ws.merge_cells(value_rng)
        h = ws[head_rng.split(":")[0]]
        v = ws[value_rng.split(":")[0]]
        h.value = label
        v.value = value
        h.font = Font(size=9, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="607D9B")
        h.alignment = Alignment(horizontal="center", vertical="center")
        v.font = Font(size=19, bold=True, color=navy)
        v.fill = PatternFill("solid", fgColor=soft)
        v.alignment = Alignment(horizontal="center", vertical="center")

    # Method
    ws.merge_cells("B12:H12")
    ws["B12"] = "CONSOLIDATION METHOD"
    ws["B12"].font = Font(size=12, bold=True, color="FFFFFF")
    ws["B12"].fill = PatternFill("solid", fgColor=blue)

    method_rows = [
        "✓ Summary figures are TOTALLED into the same original GSTR-2B portal cells.",
        "✓ Transaction data is retained at invoice level and appended across periods.",
        "✓ GSTR-2B Month is added to transaction sheets for source-period traceability.",
        "✓ Original GSTR-2B portal headers and structure are retained.",
        "✓ Repeated transaction headers are prevented from entering the data.",
    ]
    for r, text in enumerate(method_rows, 13):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r,2,text)
        ws.cell(r,2).font = Font(size=10, color="34445A")
        ws.cell(r,2).alignment = Alignment(vertical="center", wrap_text=True)

    # Status
    ws.merge_cells("B20:H20")
    ws["B20"] = "PROCESSING STATUS   ✓ COMPLETED"
    ws["B20"].font = Font(size=12, bold=True, color="FFFFFF")
    ws["B20"].fill = PatternFill("solid", fgColor=green)
    ws["B20"].alignment = Alignment(horizontal="center", vertical="center")

    ws["B22"] = "Summary sheets"
    ws["C22"] = 4
    ws["E22"] = "Transaction sheets"
    ws["F22"] = 16

    ws["B23"] = "Summary calculation"
    ws["C23"] = "TOTALLED"
    ws["E23"] = "Transaction calculation"
    ws["F23"] = "APPENDED"

    for r in (22,23):
        for c in (2,3,5,6):
            ws.cell(r,c).font = Font(size=10, bold=(c in (2,5)), color="34445A")

    # Website
    ws.merge_cells("B26:H26")
    ws["B26"] = "Created with Pushpak Kumar's GST Automation Toolkit"
    ws["B26"].font = Font(size=10, italic=True, color=grey)
    ws["B26"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B27:H27")
    ws["B27"] = "pushpakkumar.com"
    ws["B27"].font = Font(size=11, bold=True, color="2463A6", underline="single")
    ws["B27"].alignment = Alignment(horizontal="center")
    ws["B27"].hyperlink = "https://pushpakkumar.com"

    ws.freeze_panes = "B8"

def create_validation_sheet(files, summary_counts, transaction_counts, log, wb):
    ws = wb.create_sheet("Validation", 1)

    headers = ["Check", "Result", "Details"]
    for c,h in enumerate(headers,1):
        ws.cell(1,c,h)

    checks = [
        ("Source Excel files", len(files), "Files loaded from ZIP"),
        ("Summary sheets processed", sum(1 for x in summary_counts.values() if x > 0), "Summary totals built"),
        ("Transaction sheets with data", sum(1 for x in transaction_counts.values() if x > 0), "Transaction rows appended"),
        ("Total transaction rows", sum(transaction_counts.values()), "Across all 16 transaction sheets"),
        ("Files with detected month", sum(1 for n,_ in files if month_from_filename(n) != "Unknown"), "Month parsed from filename"),
    ]

    for r, row in enumerate(checks,2):
        for c,v in enumerate(row,1):
            ws.cell(r,c,v)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ================================================================
# BUILD
# ================================================================
def build_workbook(files):
    wb = Workbook()
    wb.remove(wb.active)

    summary_counts = {}
    transaction_counts = {}

    # 1. Summary totals — ONE combined row structure, figures added.
    for s in SUMMARY_SHEETS:
        count, _ = combine_summary_sheet(files, s, wb)
        summary_counts[s] = count

    # 2. Transaction sheets — append data, month tagged.
    for s in TRANSACTION_SHEETS:
        count, _ = combine_transaction_sheet(files, s, wb)
        transaction_counts[s] = count

    # 3. Read me and validation
    create_readme(files, summary_counts, transaction_counts, None, wb)
    create_validation_sheet(files, summary_counts, transaction_counts, None, wb)

    # Put Read me first, Validation second.
    wb._sheets.sort(key=lambda ws: (
        0 if ws.title == "Read me" else
        1 if ws.title == "Validation" else
        2 if ws.title in SUMMARY_SHEETS else
        3
    ))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, summary_counts, transaction_counts

# ================================================================
# STREAMLIT UI
# ================================================================
st.markdown("""
<div class="pk-hero">
  <div class="pk-eyebrow">GST AUTOMATION TOOLKIT</div>
  <div class="pk-title">GSTR-2B Consolidator</div>
  <div class="pk-desc">Consolidate monthly GSTR-2B portal files into one reliable, reconciliation-ready workbook.</div>
</div>

<div class="pk-features">
  <div class="pk-card"><b>Summary Totalisation</b><span>Monthly ITC figures are added into the same portal cells to produce consolidated totals.</span></div>
  <div class="pk-card"><b>Invoice Traceability</b><span>Every transaction carries its GSTR-2B month for source-period identification.</span></div>
  <div class="pk-card"><b>Portal Structure</b><span>Original portal headers and layout are retained rather than rebuilt.</span></div>
  <div class="pk-card"><b>Validation</b><span>Processing checks help identify missing or unexpected source data.</span></div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="pk-section">Upload GSTR-2B Files</div>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Upload ZIP containing ORIGINAL monthly GSTR-2B portal Excel files",
    type=["zip"]
)

st.info(
    "⏱️ **Processing Note:** Once submitted, it may take around **10–12 minutes** "
    "to process your GSTR-2B files. Please have a little patience and keep this page open."
)

if uploaded:
    files = []

    with zipfile.ZipFile(uploaded) as z:
        for info in z.infolist():
            if info.is_dir():
                continue

            name = Path(info.filename).name

            if name.startswith("~$"):
                continue

            if name.lower().endswith((".xlsx", ".xlsm")):
                files.append((name, z.read(info)))

    files.sort(key=lambda x: x[0])

    if not files:
        st.error("No Excel files found inside the ZIP.")
        st.stop()

    detected = pd.DataFrame({
        "Source File": [x[0] for x in files],
        "GSTR-2B Month": [month_from_filename(x[0]) for x in files]
    })

    st.success(f"Found {len(files)} original GSTR-2B Excel file(s).")
    st.dataframe(detected, use_container_width=True, hide_index=True)

    if st.button("🚀 Combine GSTR-2B", type="primary", use_container_width=True):
        progress = st.progress(0)

        with st.spinner(
            "Analysing portal structures, preserving headers, "
            "totaling summary figures and appending transactions..."
        ):
            output, summary_counts, transaction_counts = build_workbook(files)

        progress.progress(100)

        st.success("Completed successfully.")

        st.markdown('<div class="pk-section">Summary Totalisation</div>', unsafe_allow_html=True)
        st.dataframe(
            pd.DataFrame({
                "Sheet": SUMMARY_SHEETS,
                "Numeric cells added": [summary_counts[s] for s in SUMMARY_SHEETS]
            }),
            use_container_width=True,
            hide_index=True
        )

        st.markdown('<div class="pk-section">Transaction Consolidation</div>', unsafe_allow_html=True)
        st.dataframe(
            pd.DataFrame({
                "Sheet": TRANSACTION_SHEETS,
                "Rows appended": [transaction_counts[s] for s in TRANSACTION_SHEETS]
            }),
            use_container_width=True,
            hide_index=True
        )

        st.download_button(
            "⬇️ Download GSTR-2B-Combined.xlsx",
            data=output.getvalue(),
            file_name="GSTR-2B-Combined.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.markdown("""
<div class="pk-footer">
GST Automation Toolkit &nbsp;•&nbsp; GSTR-2B Consolidator
&nbsp;•&nbsp; <a href="https://pushpakkumar.com" target="_blank">pushpakkumar.com</a>
</div>
""", unsafe_allow_html=True)
